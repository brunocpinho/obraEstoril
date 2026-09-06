import openpyxl, os, sys
sys.stdout.reconfigure(encoding='utf-8')

dir_path = r'C:\Users\bruno\GLN Engenharia LTDA\Gabriel Paiva Moreira Alves - 22- Bruno e Kelly - Estoril\05- Fechamento financeiro Cliente'
target_name = [f for f in os.listdir(dir_path) if f.startswith('10 -')][0]
fpath = os.path.join(dir_path, target_name)
wb = openpyxl.load_workbook(fpath, data_only=True)
ws = wb['Total']

items = []
for r in range(7, ws.max_row + 1):
    ico = ws.cell(row=r, column=1).value
    desc = ws.cell(row=r, column=2).value
    unid = ws.cell(row=r, column=3).value
    qtd = ws.cell(row=r, column=4).value
    unit = ws.cell(row=r, column=5).value
    tot1 = ws.cell(row=r, column=6).value
    tot2 = ws.cell(row=r, column=7).value
    tot3 = ws.cell(row=r, column=8).value
    real1 = ws.cell(row=r, column=9).value
    real2 = ws.cell(row=r, column=10).value
    real3 = ws.cell(row=r, column=11).value
    pct = ws.cell(row=r, column=12).value
    mat = ws.cell(row=r, column=13).value
    mo = ws.cell(row=r, column=14).value
    ganho = ws.cell(row=r, column=15).value

    ico_str = str(ico).strip() if ico is not None else ''
    desc_str = str(desc).strip() if desc is not None else ''
    if not ico_str and not desc_str:
        continue
    
    # Determine level
    if len(ico_str) <= 2 and not '.' in ico_str:
        level = 1
        orcado = tot1
        realizado = real1
    elif ico_str.count('.') == 1:
        level = 2
        orcado = tot2
        realizado = real2
    else:
        level = 3
        orcado = tot3
        realizado = real3

    items.append({
        'row': r,
        'level': level,
        'ico': ico_str,
        'desc': desc_str,
        'unid': unid,
        'qtd': qtd,
        'unit': unit,
        'orcado': orcado or 0,
        'realizado': realizado or 0,
        'pct': pct,
        'mat': mat or 0,
        'mo': mo or 0,
        'ganho': ganho
    })

print(f'Total items extracted: {len(items)}')

print('\n' + '='*80)
print('=== NÍVEL 1 (MACROETAPAS) ===')
print('='*80)
for it in items:
    if it['level'] == 1:
        pct_val = it['pct']
        if isinstance(pct_val, (int, float)):
            p_str = f"{pct_val*100:6.2f}%"
        else:
            p_str = f"{str(pct_val):>7}"
        print(f"ICO {it['ico']:<4} | {it['desc']:<35} | Orçado: R$ {it['orcado']:>11,.2f} | Realizado: R$ {it['realizado']:>10,.2f} | {p_str}")

print('\n' + '='*80)
print('=== NÍVEL 2 (SUBETAPAS) ===')
print('='*80)
for it in items:
    if it['level'] == 2:
        pct_val = it['pct']
        if isinstance(pct_val, (int, float)):
            p_str = f"{pct_val*100:6.2f}%"
        else:
            p_str = f"{str(pct_val):>7}"
        print(f"  {it['ico']:<8} | {it['desc']:<35} | Orçado: R$ {it['orcado']:>11,.2f} | Realizado: R$ {it['realizado']:>10,.2f} | {p_str}")

print('\n' + '='*80)
print('=== ITENS COM CUSTO REALIZADO > 0 (TODOS OS NÍVEIS 3 DETALHADOS) ===')
print('='*80)
for it in items:
    if it['level'] == 3 and it['realizado'] > 0:
        pct_val = it['pct']
        if isinstance(pct_val, (int, float)):
            p_str = f"{pct_val*100:6.2f}%"
        else:
            p_str = f"{str(pct_val):>7}"
        print(f"    {it['ico']:<10} | {it['desc']:<40} | Orçado: R$ {it['orcado']:>10,.2f} | Real: R$ {it['realizado']:>10,.2f} | {p_str} | Mat: R$ {it['mat']:>9,.2f} | MO: R$ {it['mo']:>9,.2f}")
